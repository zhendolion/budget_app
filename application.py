from flask import current_app, Flask, render_template, request, redirect, url_for, flash, jsonify, send_file, Response
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import extract, func,case
from sqlalchemy.exc import SQLAlchemyError
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.exceptions import BadRequest
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.cron import CronTrigger
from openpyxl import Workbook
from datetime import datetime
from openpyxl import load_workbook
import pandas as pd
import plotly
import plotly.express as px
import plotly.graph_objects as go
import json
import re #For email validation 
import os
import io
import csv
import secrets


app = Flask(__name__, static_folder='static')
secret_key = secrets.token_urlsafe(16)
app.secret_key = secret_key  # Replace with a secure secret key in production

# Initialize Flask-Login
# Initialize Flask-Login
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'auth_login'
login_manager.login_message = 'Please log in to access this page.'
login_manager.login_message_category = 'info'

# User loader callback
@login_manager.user_loader
def load_user(user_id):
    # Update to use Session.get() instead of Query.get()
    return db.session.get(User, int(user_id))


# Database Configuration
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///budget.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db = SQLAlchemy(app)

#
# Models
class Category(db.Model):
    __tablename__ = 'category'
    
    
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(50), nullable=False)
    type = db.Column(db.String(20), nullable=False)  # 'income' or 'expense'
    description = db.Column(db.String(200))
    budget_limit = db.Column(db.Float, default=0)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)

    __table_args__ = (
        db.UniqueConstraint('user_id', 'name', 'type', name='unique_category_per_user'),
    )
    @property
    def transaction_count(self):
        return Transaction.query.filter_by(category_id=self.id).count()

    @classmethod
    def get_user_categories(cls, user_id, category_type=None):
        query = cls.query.filter_by(user_id=user_id)
        if category_type:
            query = query.filter_by(type=category_type)
        return query.order_by(cls.name).all()
    
    
    def __repr__(self):
        return f'<Category {self.name}>'


class Transaction(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    date = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)
    amount = db.Column(db.Float, nullable=False)
    description = db.Column(db.String(200))
    type = db.Column(db.String(20), nullable=False)  # 'Income' or 'Expense'
    category_id = db.Column(db.Integer, db.ForeignKey('category.id'), nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    
    # Add the relationship to Category
    category = db.relationship('Category', backref='transactions', lazy=True)
    user = db.relationship('User', backref='transactions')

    def __repr__(self):
        return f'<Transaction {self.description}: {self.amount}>'

class BudgetGoal(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    category_id = db.Column(db.Integer, db.ForeignKey('category.id'), nullable=False)
    amount = db.Column(db.Float, nullable=False)
    period = db.Column(db.String(10), nullable=False)  # daily, weekly, monthly, quarterly, yearly
    recurring = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    # Relationships
    user = db.relationship('User', backref='budget_goals')
    category = db.relationship('Category', backref='budget_goals')

    
    def __repr__(self):
        return f'<BudgetGoal {self.category.name}: {self.amount}>'

    
class RecurringTransaction(db.Model):
    __tablename__ = 'recurring_transaction'

    id = db.Column(db.Integer, primary_key=True)
    type = db.Column(db.String(20), nullable=False)  # 'income' or 'expense'
    category = db.Column(db.String(50), nullable=False)
    amount = db.Column(db.Float, nullable=False)
    frequency = db.Column(db.String(20), nullable=False)  # 'daily', 'weekly', 'monthly', 'yearly'
    description = db.Column(db.String(200))
    start_date = db.Column(db.DateTime, nullable=False)
    end_date = db.Column(db.DateTime)
    last_processed = db.Column(db.DateTime)
    is_active = db.Column(db.Boolean, default=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    
    def __repr__(self):
        return f'<RecurringTransaction {self.description}: {self.amount}>'

class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    email = db.Column(db.String(120), unique=True, nullable=False)
    first_name = db.Column(db.String(64))
    last_name = db.Column(db.String(64))
    password_hash = db.Column(db.String(128))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    # Relationships
    categories = db.relationship('Category', backref='user', lazy=True)
    recurring_transactions = db.relationship('RecurringTransaction', backref='user', lazy=True)
    
    def set_password(self, password):
        self.password_hash = generate_password_hash(password)

    def check_password(self, password):
        return check_password_hash(self.password_hash, password)

    def get_full_name(self):
        """Return user's full name or email if name not set"""
        if self.first_name and self.last_name:
            return f"{self.first_name} {self.last_name}"
        return self.email

    def __repr__(self):
        return f'<User {self.email}>'





# Helper Functions
def init_db():
    with app.app_context():
        db.create_all()
        
        # Check if system categories exist specifically for user_id=1
        system_categories = Category.query.filter_by(user_id=1).first()
        
        if not system_categories:
            # Create default categories for the system user
            default_categories = [
                # Income categories
                Category(
                    name='Salary',
                    type='Income',
                    user_id=1,
                    description='System default - Salary'
                ),
                Category(
                    name='Freelance',
                    type='Income',
                    user_id=1,
                    description='System default - Freelance'
                ),
                Category(
                    name='Investments',
                    type='Income',
                    user_id=1,
                    description='System default - Investments'
                ),
                Category(
                    name='Other Income',
                    type='Income',
                    user_id=1,
                    description='System default - Other Income'
                ),
                
                # Expense categories
                Category(
                    name='Groceries',
                    type='Expense',
                    user_id=1,
                    description='System default - Groceries'
                ),
                Category(
                    name='Rent',
                    type='Expense',
                    user_id=1,
                    description='System default - Rent'
                ),
                Category(
                    name='Utilities',
                    type='Expense',
                    user_id=1,
                    description='System default - Utilities'
                ),
                Category(
                    name='Transportation',
                    type='Expense',
                    user_id=1,
                    description='System default - Transportation'
                ),
                Category(
                    name='Entertainment',
                    type='Expense',
                    user_id=1,
                    description='System default - Entertainment'
                ),
                Category(
                    name='Healthcare',
                    type='Expense',
                    user_id=1,
                    description='System default - Healthcare'
                ),
                Category(
                    name='Shopping',
                    type='Expense',
                    user_id=1,
                    description='System default - Shopping'
                ),
                Category(
                    name='Restaurants',
                    type='Expense',
                    user_id=1,
                    description='System default - Restaurants'
                )
            ]
            
            try:
                db.session.add_all(default_categories)
                db.session.commit()
                print("System default categories added successfully!")
            except Exception as e:
                db.session.rollback()
                print(f"Error adding system default categories: {str(e)}")
        else:
            print("System categories already exist, skipping initialization.")


def calculate_current_spending():
    expenses = Transaction.query.filter_by(type='Expense').all()
    spending = {}
    for expense in expenses:
        if expense.category in spending:
            spending[expense.category] += expense.amount
        else:
            spending[expense.category] = expense.amount
    return spending

def get_all_categories(user_id=None, category_type=None):
    query = Category.query
    
    if user_id:
        query = query.filter_by(user_id=user_id)
    
    if category_type:
        query = query.filter_by(type=category_type)
        
    return query.order_by(Category.type, Category.name).all()

# Then you can use it like:
# get_all_categories(user_id=current_user.id, category_type='Expense')
# get_all_categories(user_id=current_user.id, category_type='Income')


def get_budget_categories():
    """Get budget categories with their spending progress for the current user"""
    try:
        # Get the current month's date range
        today = datetime.now()
        start_of_month = datetime(today.year, today.month, 1)
        end_of_month = (start_of_month + timedelta(days=32)).replace(day=1) - timedelta(days=1)

        # Get all categories for the current user
        categories = Category.query.filter_by(user_id=current_user.id).all()
        budget_progress = []

        for category in categories:
            # Get total spent in this category for the current month
            total_spent = db.session.query(func.sum(Transaction.amount)).\
                filter(Transaction.category_id == category.id,
                       Transaction.type == 'Expense',
                       Transaction.date >= start_of_month,
                       Transaction.date <= end_of_month,
                       Transaction.user_id == current_user.id).scalar() or 0

            # Calculate percentage of budget spent
            if category.budget_limit:
                percentage = (total_spent / category.budget_limit) * 100
            else:
                percentage = 0

            budget_progress.append({
                'name': category.name,
                'spent': total_spent,
                'budget': category.budget_limit or 0,
                'percentage': round(percentage, 1)
            })

        return budget_progress

    except Exception as e:
        print(f"Error in get_budget_categories: {str(e)}")
        return []
    
def process_recurring_transactions():
    with app.app_context():
        try:
            now = datetime.now()
            recurring_transactions = RecurringTransaction.query.filter_by(is_active=True).all()
            
            for recurring in recurring_transactions:
                try:
                    # Skip if end date is set and passed
                    if recurring.end_date and recurring.end_date < now:
                        continue
                        
                    # Get the last processed date or use start date if never processed
                    last_processed = recurring.last_processed or recurring.start_date
                    next_date = None
                    
                    # Calculate the next date based on frequency
                    if recurring.frequency == 'daily':
                        next_date = last_processed + timedelta(days=1)
                    elif recurring.frequency == 'weekly':
                        next_date = last_processed + timedelta(weeks=1)
                    elif recurring.frequency == 'monthly':
                        next_date = last_processed + relativedelta(months=1)
                    elif recurring.frequency == 'yearly':
                        next_date = last_processed + relativedelta(years=1)
                    
                    # Process all pending transactions up to current date
                    while next_date <= now:
                        try:
                            # Get the category ID based on the category name
                            category = Category.query.filter_by(
                                name=recurring.category,
                                user_id=recurring.user_id
                            ).first()
                            
                            if not category:
                                current_app.logger.error(f"Category not found for recurring transaction: {recurring.id}")
                                break

                            # Create new transaction
                            new_transaction = Transaction(
                                date=next_date,
                                type=recurring.type,
                                category_id=category.id,
                                amount=recurring.amount,
                                description=f"[Recurring] {recurring.description}" if recurring.description else "[Recurring Transaction]",
                                user_id=recurring.user_id
                            )
                            
                            db.session.add(new_transaction)
                            recurring.last_processed = next_date
                            
                            # Calculate next date for the loop
                            if recurring.frequency == 'daily':
                                next_date += timedelta(days=1)
                            elif recurring.frequency == 'weekly':
                                next_date += timedelta(weeks=1)
                            elif recurring.frequency == 'monthly':
                                next_date += relativedelta(months=1)
                            elif recurring.frequency == 'yearly':
                                next_date += relativedelta(years=1)
                                
                        except Exception as e:
                            current_app.logger.error(f"Error processing recurring transaction {recurring.id}: {str(e)}")
                            db.session.rollback()
                            break
                            
                except Exception as e:
                    current_app.logger.error(f"Error processing recurring transaction {recurring.id}: {str(e)}")
                    continue
            
            try:
                db.session.commit()
                current_app.logger.info("Recurring transactions processed successfully")
            except Exception as e:
                current_app.logger.error(f"Error committing recurring transactions: {str(e)}")
                db.session.rollback()
                
        except Exception as e:
            current_app.logger.error(f"Error in process_recurring_transactions: {str(e)}")

def process_pending_recurring_transactions():
    """Process any pending recurring transactions immediately"""
    try:
        now = datetime.now()
        recurring_transactions = RecurringTransaction.query.filter_by(is_active=True).all()
        
        needs_processing = False
        for recurring in recurring_transactions:
            if recurring.end_date and recurring.end_date < now:
                continue
                
            last_processed = recurring.last_processed or recurring.start_date
            if last_processed < now:
                needs_processing = True
                break
        
        if needs_processing:
            process_recurring_transactions()
            current_app.logger.info("Pending recurring transactions processed")
        else:
            current_app.logger.info("No pending recurring transactions to process")
            
    except Exception as e:
        current_app.logger.error(f"Error checking recurring transactions: {str(e)}")


def get_budget_goals():
    goals = {}
    for goal in BudgetGoal.query.all():
        goals[goal.category] = goal.amount
    return goals

@app.context_processor
def utility_processor():
    def format_currency(amount):
        if amount is None or not isinstance(amount, (int, float)):
            return "0 Ft"
        try:
            # Convert to float and get absolute value for formatting
            amount = float(amount)
            
            # Format with thousand separator and no decimal places
            formatted = "{:,.0f}".format(abs(amount))
            
            # Replace comma with space for Hungarian number formatting
            formatted = formatted.replace(",", " ")
            
            # Add negative sign if amount is negative
            if amount < 0:
                return f"-{formatted} Ft"
            return f"{formatted} Ft"
            
        except (ValueError, TypeError):
            return "0 Ft"

    return dict(
        format_currency=format_currency,
        min=min
    )


# Routes

app = Flask(__name__)
logging.basicConfig(level=logging.DEBUG)

@app.route('/health')
def health_check():
    app.logger.info("Health check endpoint hit")
    return 'OK', 200

@app.route('/')
@login_required
def home():
    try:
        # Calculate date range for current month
        today = datetime.now()
        start_of_month = datetime(today.year, today.month, 1)
        end_of_month = start_of_month + relativedelta(months=1)

        # Query transactions for current month
        monthly_transactions = db.session.query(Transaction).join(
            Category
        ).filter(
            Transaction.user_id == current_user.id,
            Transaction.date >= start_of_month,
            Transaction.date < end_of_month
        ).all()

        # Calculate monthly totals
        monthly_income = sum(t.amount for t in monthly_transactions if t.type.lower() == 'income')
        monthly_expenses = sum(t.amount for t in monthly_transactions if t.type.lower() == 'expense')
        total_balance = monthly_income - monthly_expenses

        # Get recent transactions with proper relationship loading
        recent_transactions = db.session.query(Transaction).join(
            Category
        ).filter(
            Transaction.user_id == current_user.id
        ).order_by(
            Transaction.date.desc()
        ).limit(5).all()

        # Get budget categories and their spending
        budget_categories = []
        expense_categories = Category.query.filter(
            Category.user_id == current_user.id,
            func.lower(Category.type) == 'expense'
        ).all()

        for category in expense_categories:
            # Calculate spent amount for this category in current month
            spent = sum(t.amount for t in monthly_transactions 
                       if t.category_id == category.id and t.type.lower() == 'expense')

            # Calculate percentage of budget used
            if category.budget_limit and category.budget_limit > 0:
                percentage = (spent / category.budget_limit) * 100
            else:
                percentage = 0

            budget_categories.append({
                'name': category.name,
                'spent': spent,
                'budget_limit': category.budget_limit,
                'percentage': min(round(percentage, 1), 100)
            })

        current_app.logger.info(f"Successfully loaded dashboard for user {current_user.id}")
        
        return render_template('index.html',
                             total_balance=total_balance,
                             monthly_income=monthly_income,
                             monthly_expenses=monthly_expenses,
                             recent_transactions=recent_transactions,
                             budget_categories=budget_categories,
                             current_month=today.strftime('%B %Y'))

    except Exception as e:
        current_app.logger.error(f"Error in home route: {str(e)}", exc_info=True)
        flash('An error occurred while loading the dashboard', 'error')
        return render_template('index.html',
                             total_balance=0,
                             monthly_income=0,
                             monthly_expenses=0,
                             recent_transactions=[],
                             budget_categories=[],
                             current_month=datetime.now().strftime('%B %Y'))


@app.route('/add_transaction', methods=['GET', 'POST'])
@login_required
def add_transaction():
    if request.method == 'POST':
        try:
            # Get form data
            date_str = request.form.get('date')
            amount = float(request.form.get('amount'))
            description = request.form.get('description')
            transaction_type = request.form.get('type').lower()
            category_id = request.form.get('category')

            # Validate inputs
            if not all([date_str, amount, transaction_type, category_id]):
                flash('All required fields must be filled out', 'error')
                return redirect(url_for('add_transaction'))

            # Convert date string to datetime object
            date = datetime.strptime(date_str, '%Y-%m-%d')

            # Create new transaction
            new_transaction = Transaction(
                date=date,
                amount=amount,
                description=description,
                type=transaction_type,
                category_id=category_id,
                user_id=current_user.id
            )

            db.session.add(new_transaction)
            
            # Update budget goal progress if it's an expense
            if transaction_type == 'expense':
                budget_goal = BudgetGoal.query.filter_by(category_id=category_id).first()
                if budget_goal:
                    current_app.logger.info(f"""
                        New expense added for budget goal:
                        Category: {new_transaction.category.name}
                        Amount: {amount}
                        Date: {date}
                        Budget Period: {budget_goal.period}
                    """)

            db.session.commit()

            flash('Transaction added successfully!', 'success')
            return redirect(url_for('view_transactions'))

        except ValueError:
            db.session.rollback()
            flash('Invalid amount format', 'error')
            return redirect(url_for('add_transaction'))
        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f"Error adding transaction: {str(e)}")
            flash(f'Error adding transaction: {str(e)}', 'error')
            return redirect(url_for('add_transaction'))

    # GET request - show form
    try:
        # Get income and expense categories separately
        income_categories = Category.query.filter_by(
            user_id=current_user.id,
            type='Income'
        ).all()
        
        expense_categories = Category.query.filter_by(
            user_id=current_user.id,
            type='Expense'
        ).all()

        return render_template('add_transaction.html',
                             income_categories=income_categories,
                             expense_categories=expense_categories)
    except Exception as e:
        current_app.logger.error(f"Error loading categories: {str(e)}")
        flash('Error loading categories', 'error')
        return redirect(url_for('home'))

@app.route('/view_transactions')
@login_required
def view_transactions():
    try:
        # Process any pending recurring transactions first
        process_pending_recurring_transactions()
        
        # Get filter parameters
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        category_id = request.args.get('category_id')
        transaction_type = request.args.get('type')

        # Base query
        query = Transaction.query.filter_by(user_id=current_user.id)

        # Apply filters
        if start_date:
            query = query.filter(Transaction.date >= datetime.strptime(start_date, '%Y-%m-%d'))
        if end_date:
            query = query.filter(Transaction.date <= datetime.strptime(end_date, '%Y-%m-%d'))
        if category_id:
            query = query.filter(Transaction.category_id == category_id)
        if transaction_type:
            query = query.filter(Transaction.type == transaction_type)

        # Order by date
        transactions = query.order_by(Transaction.date.desc()).all()

        # Get categories for filter dropdown
        categories = Category.query.filter_by(user_id=current_user.id).all()

        # Calculate totals
        total_income = sum(t.amount for t in transactions if t.type.lower() == 'income')
        total_expenses = sum(t.amount for t in transactions if t.type.lower() == 'expense')
        net_amount = total_income - total_expenses

        # Get budget goals progress
        budget_goals = []
        current_date = datetime.now()
        goals = BudgetGoal.query.join(Category).filter(Category.user_id == current_user.id).all()
        
        for goal in goals:
            # Calculate spent amount based on period
            query = db.session.query(func.sum(Transaction.amount)).filter(
                Transaction.category_id == goal.category_id,
                Transaction.type == 'expense',
                Transaction.user_id == current_user.id
            )

            if goal.period == 'monthly':
                query = query.filter(
                    extract('month', Transaction.date) == current_date.month,
                    extract('year', Transaction.date) == current_date.year
                )
            else:  # yearly
                query = query.filter(
                    extract('year', Transaction.date) == current_date.year
                )

            spent = query.scalar() or 0
            percentage = (spent / goal.amount * 100) if goal.amount > 0 else 0

            budget_goals.append({
                'category': goal.category.name,
                'spent': round(spent, 2),
                'amount': goal.amount,
                'percentage': round(percentage, 1)
            })

        return render_template('view_transactions.html',
                             transactions=transactions,
                             categories=categories,
                             total_income=total_income,
                             total_expenses=total_expenses,
                             net_amount=net_amount,
                             budget_goals=budget_goals)

    except Exception as e:
        current_app.logger.error(f"Error viewing transactions: {str(e)}")
        flash('Error loading transactions', 'error')
        return redirect(url_for('home'))


@app.route('/update_transaction/<int:transaction_id>', methods=['POST'])
@login_required  # Add login protection if not already present
def update_transaction(transaction_id):
    try:
        transaction = Transaction.query.get_or_404(transaction_id)
        
        # Validate form data
        if not all(key in request.form for key in ['type', 'category_id', 'amount', 'date', 'description']):
            flash('Missing required fields', 'error')
            return redirect(url_for('view_transactions'))
        
        try:
            amount = float(request.form['amount'])
            if amount <= 0:
                raise ValueError("Amount must be greater than 0")
        except ValueError as e:
            flash(f'Invalid amount: {str(e)}', 'error')
            return redirect(url_for('view_transactions'))

        # Update transaction
        transaction.type = request.form['type']
        transaction.category_id = request.form['category_id']
        transaction.amount = amount
        transaction.description = request.form['description']
        transaction.date = datetime.strptime(request.form['date'], '%Y-%m-%d')
        
        db.session.commit()
        flash('Transaction updated successfully', 'success')
    except ValueError as e:
        db.session.rollback()
        current_app.logger.error(f"Value error updating transaction: {str(e)}")
        flash(f'Error updating transaction: {str(e)}', 'error')
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error updating transaction: {str(e)}")
        flash('An error occurred while updating the transaction', 'error')
    
    return redirect(url_for('view_transactions'))

@app.route('/set_budget_goals', methods=['GET', 'POST'])
@login_required
def set_budget_goals():
    try:
        if request.method == 'POST':
            # Get form data
            category_id = request.form.get('category')
            amount = request.form.get('amount')
            period = request.form.get('period')
            recurring = request.form.get('recurring', 'false') == 'true'

            # Validate inputs
            if not all([category_id, amount, period]):
                flash('All fields are required', 'error')
                return redirect(url_for('set_budget_goals'))

            try:
                amount = float(amount)
                if amount <= 0:
                    raise ValueError
            except ValueError:
                flash('Amount must be a positive number', 'error')
                return redirect(url_for('set_budget_goals'))

            valid_periods = ['daily', 'weekly', 'monthly', 'quarterly', 'yearly']
            if period not in valid_periods:
                flash('Invalid period specified', 'error')
                return redirect(url_for('set_budget_goals'))

            # Check if budget goal already exists for this category and period
            existing_goal = BudgetGoal.query.filter_by(
                category_id=category_id,
                period=period,
                user_id=current_user.id
            ).first()

            if existing_goal:
                flash('A budget goal already exists for this category and period', 'error')
                return redirect(url_for('set_budget_goals'))

            # Create new budget goal
            new_goal = BudgetGoal(
                category_id=category_id,
                amount=amount,
                period=period,
                recurring=recurring,
                user_id=current_user.id
            )

            db.session.add(new_goal)
            db.session.commit()
            flash('Budget goal set successfully!', 'success')
            return redirect(url_for('set_budget_goals'))

        # GET request - show the form
        categories = Category.query.filter_by(user_id=current_user.id).all()
        current_app.logger.info(f"Found {len(categories)} categories for user {current_user.id}")

        # Get all budget goals for the current user
        goals = BudgetGoal.query.filter_by(user_id=current_user.id).all()
        current_app.logger.info(f"Found {len(goals)} budget goals for user {current_user.id}")

        # Get recurring transactions
        recurring_transactions = RecurringTransaction.query.filter_by(user_id=current_user.id).all()
        current_app.logger.info(f"Found {len(recurring_transactions)} recurring transactions for user {current_user.id}")

        budget_goals = []
        current_date = datetime.now()

        for goal in goals:
            try:
                # Calculate spent amount based on period
                query = db.session.query(func.sum(Transaction.amount)).\
                    filter(
                        Transaction.category_id == goal.category_id,
                        Transaction.type == 'expense',
                        Transaction.user_id == current_user.id
                    )

                # Adjust date filter based on period
                if goal.period == 'daily':
                    query = query.filter(
                        func.date(Transaction.date) == current_date.date()
                    )
                elif goal.period == 'weekly':
                    start_of_week = current_date - timedelta(days=current_date.weekday())
                    query = query.filter(
                        Transaction.date >= start_of_week,
                        Transaction.date < start_of_week + timedelta(days=7)
                    )
                elif goal.period == 'monthly':
                    query = query.filter(
                        extract('month', Transaction.date) == current_date.month,
                        extract('year', Transaction.date) == current_date.year
                    )
                elif goal.period == 'quarterly':
                    current_quarter = (current_date.month - 1) // 3 + 1
                    quarter_start = datetime(current_date.year, (current_quarter - 1) * 3 + 1, 1)
                    quarter_end = quarter_start + relativedelta(months=3)
                    query = query.filter(
                        Transaction.date >= quarter_start,
                        Transaction.date < quarter_end
                    )
                elif goal.period == 'yearly':
                    query = query.filter(
                        extract('year', Transaction.date) == current_date.year
                    )

                spent = query.scalar() or 0
                percentage = (spent / goal.amount) * 100 if goal.amount > 0 else 0

                category = Category.query.get(goal.category_id)
                budget_goals.append({
                    'id': goal.id,
                    'category': category.name,
                    'category_id': category.id,
                    'amount': goal.amount,
                    'spent': spent,
                    'percentage': percentage,
                    'period': goal.period
                })

            except Exception as e:
                current_app.logger.error(f"Error processing budget goal {goal.id}: {str(e)}")
                continue

        return render_template('budget_goals.html',
                             categories=categories,
                             budget_goals=budget_goals,
                             recurring_transactions=recurring_transactions)

    except Exception as e:
        current_app.logger.error(f"Error in set_budget_goals: {str(e)}", exc_info=True)
        flash('An error occurred while loading the page', 'error')
        return redirect(url_for('home'))




@app.route('/edit_budget_goal/<int:goal_id>', methods=['POST'])
@login_required
def edit_budget_goal(goal_id):
    try:
        goal = BudgetGoal.query.get_or_404(goal_id)
        
        # Get form data
        amount = float(request.form.get('amount', 0))
        period = request.form.get('period')

        # Validate inputs
        if amount <= 0:
            flash('Amount must be greater than 0', 'error')
            return redirect(url_for('set_budget_goals'))

        if period not in ['monthly', 'yearly']:
            flash('Invalid period specified', 'error')
            return redirect(url_for('set_budget_goals'))

        # Update the goal
        goal.amount = amount
        goal.period = period
        
        db.session.commit()
        flash('Budget goal updated successfully!', 'success')
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error updating budget goal: {str(e)}")
        flash('Error updating budget goal', 'error')
    
    return redirect(url_for('set_budget_goals'))

@app.route('/delete_budget_goal/<int:goal_id>', methods=['POST'])
@login_required
def delete_budget_goal(goal_id):
    try:
        goal = BudgetGoal.query.get_or_404(goal_id)
        db.session.delete(goal)
        db.session.commit()
        flash('Budget goal deleted successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error deleting budget goal: {str(e)}")
        flash('Error deleting budget goal', 'error')
    
    return redirect(url_for('set_budget_goals'))

@app.route('/manage-categories', methods=['GET', 'POST'])
@login_required
def manage_categories():
    try:
        if request.method == 'POST':
            action = request.form.get('action')
            
            if action == 'add':
                name = request.form.get('name', '').strip()
                category_type = request.form.get('type', '').capitalize()
                
                # Check if category already exists for this user
                existing_category = Category.query.filter(
                    db.func.lower(Category.name) == name.lower(),  # Case-insensitive comparison
                    Category.type == category_type,
                    Category.user_id == current_user.id
                ).first()
                
                if existing_category:
                    flash(f'A category named "{name}" already exists for {category_type}!', 'error')
                else:
                    new_category = Category(
                        name=name,
                        type=category_type,
                        user_id=current_user.id
                    )
                    db.session.add(new_category)
                    db.session.commit()
                    flash('Category added successfully!', 'success')
                
            elif action == 'edit':
                category_id = request.form.get('category_id')
                name = request.form.get('name', '').strip()
                category_type = request.form.get('type', '').capitalize()
                
                category = Category.query.filter_by(
                    id=category_id,
                    user_id=current_user.id
                ).first()
                
                if category:
                    # Check if another category with the same name exists
                    existing_category = Category.query.filter(
                        db.func.lower(Category.name) == name.lower(),  # Case-insensitive comparison
                        Category.type == category_type,
                        Category.user_id == current_user.id,
                        Category.id != category_id  # Exclude current category
                    ).first()
                    
                    if existing_category:
                        flash(f'A category named "{name}" already exists for {category_type}!', 'error')
                    else:
                        category.name = name
                        category.type = category_type
                        db.session.commit()
                        flash('Category updated successfully!', 'success')
                else:
                    flash('Category not found.', 'error')
                    
            elif action == 'delete':
                category_id = request.form.get('category_id')
                
                # Check if category has associated transactions
                transactions_exist = Transaction.query.filter_by(
                    category_id=category_id
                ).first() is not None
                
                if transactions_exist:
                    flash('Cannot delete category with existing transactions.', 'error')
                    return redirect(url_for('manage_categories'))
                
                category = Category.query.filter_by(
                    id=category_id,
                    user_id=current_user.id
                ).first()
                
                if category:
                    db.session.delete(category)
                    db.session.commit()
                    flash('Category deleted successfully!', 'success')
                else:
                    flash('Category not found.', 'error')

        # Get categories using your get_all_categories function
        expense_categories = get_all_categories(
            user_id=current_user.id, 
            category_type='Expense'
        )
        
        income_categories = get_all_categories(
            user_id=current_user.id, 
            category_type='Income'
        )

        return render_template('manage_categories.html',
                             expense_categories=expense_categories,
                             income_categories=income_categories)

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error in manage_categories: {str(e)}")
        flash('An error occurred while managing categories', 'error')
        return redirect(url_for('manage_categories'))


@app.route('/add-recurring-transaction', methods=['GET', 'POST'])
@login_required
def add_recurring_transaction():
    if request.method == 'POST':
        try:
            # Log the incoming form data
            current_app.logger.debug(f"Form data: {request.form}")
            
            # Get form data
            category_name = request.form.get('category')
            
            # Create new recurring transaction
            new_recurring = RecurringTransaction(
                type=request.form.get('type'),
                category=category_name,
                amount=float(request.form.get('amount')),
                frequency=request.form.get('frequency'),
                description=request.form.get('description'),
                start_date=datetime.strptime(request.form.get('start_date'), '%Y-%m-%d'),
                end_date=datetime.strptime(request.form.get('end_date'), '%Y-%m-%d') if request.form.get('end_date') else None,
                is_active=True,
                user_id=current_user.id
            )
            
            current_app.logger.debug(f"Created new recurring transaction: {new_recurring}")
            
            db.session.add(new_recurring)
            db.session.commit()
            
            # Verify the transaction was saved
            saved_transaction = RecurringTransaction.query.get(new_recurring.id)
            current_app.logger.info(f"Successfully saved recurring transaction: {saved_transaction}")
            
            flash('Recurring transaction added successfully!', 'success')
            return redirect(url_for('set_budget_goals'))
            
        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f"Error adding recurring transaction: {str(e)}", exc_info=True)
            flash('Error adding recurring transaction', 'error')
            return redirect(url_for('set_budget_goals'))
    
    return redirect(url_for('set_budget_goals'))
@app.route('/delete-recurring-transaction/<int:transaction_id>', methods=['POST'])
@login_required
def delete_recurring_transaction(transaction_id):
    try:
        # Find the recurring transaction
        recurring_transaction = RecurringTransaction.query.filter_by(
            id=transaction_id,
            user_id=current_user.id
        ).first()
        
        if not recurring_transaction:
            flash('Recurring transaction not found.', 'error')
            return redirect(url_for('set_budget_goals'))
        
        # Store information for confirmation message
        description = recurring_transaction.description
        
        # Delete the recurring transaction
        db.session.delete(recurring_transaction)
        db.session.commit()
        
        flash(f'Recurring transaction "{description}" has been deleted.', 'success')
        
    except Exception as e:
        current_app.logger.error(f"Error deleting recurring transaction: {str(e)}")
        db.session.rollback()
        flash('Error deleting recurring transaction.', 'error')
    
    return redirect(url_for('set_budget_goals'))

@app.route('/edit-recurring-transaction/<int:transaction_id>', methods=['POST'])
@login_required
def edit_recurring_transaction(transaction_id):
    try:
        # Find the recurring transaction
        recurring_transaction = RecurringTransaction.query.filter_by(
            id=transaction_id,
            user_id=current_user.id
        ).first()
        
        if not recurring_transaction:
            flash('Recurring transaction not found.', 'error')
            return redirect(url_for('set_budget_goals'))
        
        # Update the transaction with form data
        recurring_transaction.type = request.form.get('type')
        recurring_transaction.category = request.form.get('category')
        recurring_transaction.amount = float(request.form.get('amount'))
        recurring_transaction.frequency = request.form.get('frequency')
        recurring_transaction.description = request.form.get('description')
        recurring_transaction.start_date = datetime.strptime(request.form.get('start_date'), '%Y-%m-%d')
        
        if request.form.get('end_date'):
            recurring_transaction.end_date = datetime.strptime(request.form.get('end_date'), '%Y-%m-%d')
        else:
            recurring_transaction.end_date = None
            
        recurring_transaction.is_active = request.form.get('is_active') == 'on'
        
        db.session.commit()
        flash('Recurring transaction updated successfully!', 'success')
        
    except Exception as e:
        current_app.logger.error(f"Error updating recurring transaction: {str(e)}")
        db.session.rollback()
        flash('Error updating recurring transaction.', 'error')
    
    return redirect(url_for('set_budget_goals'))

@app.route('/account_settings', methods=['GET', 'POST'])
@login_required
def account_settings():
    try:
        if request.method == 'POST':
            if 'file' in request.files:  # Handle CSV import
                file = request.files['file']
                if file.filename == '':
                    flash('No file selected', 'error')
                    return redirect(url_for('account_settings'))
                
                if not file.filename.endswith('.csv'):
                    flash('Only CSV files are allowed', 'error')
                    return redirect(url_for('account_settings'))
                
                try:
                    # Read CSV file
                    df = pd.read_csv(file, encoding='utf-8')
                    
                    # Validate CSV structure
                    required_columns = ['date', 'amount', 'type', 'category', 'description']
                    if not all(col in df.columns for col in required_columns):
                        flash('Invalid CSV format. Required columns: date, amount, type, category, description', 'error')
                        return redirect(url_for('account_settings'))
                    
                    success_count = 0
                    error_count = 0
                    
                    for _, row in df.iterrows():
                        try:
                            # Convert date string to datetime
                            transaction_date = datetime.strptime(str(row['date']), '%Y-%m-%d')
                            
                            # Get or create category
                            category_name = str(row['category']).strip()
                            transaction_type = str(row['type']).lower()
                            
                            category = Category.query.filter_by(
                                name=category_name,
                                user_id=current_user.id
                            ).first()
                            
                            if not category:
                                category = Category(
                                    name=category_name,
                                    type=transaction_type,
                                    user_id=current_user.id
                                )
                                db.session.add(category)
                                db.session.flush()
                            
                            # Create transaction
                            transaction = Transaction(
                                date=transaction_date,
                                amount=float(row['amount']),
                                description=str(row['description']),
                                type=transaction_type,
                                category_id=category.id,
                                user_id=current_user.id
                            )
                            
                            db.session.add(transaction)
                            success_count += 1
                            
                        except Exception as row_error:
                            error_count += 1
                            current_app.logger.error(f"Error processing row: {row_error}")
                            continue
                    
                    db.session.commit()
                    flash(f'Successfully imported {success_count} transactions. {error_count} errors.', 
                          'success' if error_count == 0 else 'warning')
                    
                except Exception as e:
                    db.session.rollback()
                    current_app.logger.error(f"Error processing CSV: {str(e)}")
                    flash('Error processing CSV file', 'error')

        # Get statistics for display
        stats = {
            'total_transactions': Transaction.query.filter_by(user_id=current_user.id).count(),
            'total_categories': Category.query.filter_by(user_id=current_user.id).count(),
            'total_budget_goals': Category.query.filter(
                Category.user_id == current_user.id,
                Category.budget_limit > 0
            ).count()
        }

        return render_template('account_settings.html', user=current_user, stats=stats)

    except Exception as e:
        current_app.logger.error(f"Error in account settings: {str(e)}")
        flash('An error occurred while loading settings', 'error')
        return redirect(url_for('home'))

@app.route('/export_transactions_csv')
@login_required
def export_transactions_csv():
    try:
        # Get all transactions for the user
        transactions = Transaction.query.filter_by(user_id=current_user.id).all()
        
        # Create CSV data
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write header
        writer.writerow(['date', 'amount', 'type', 'category', 'description'])
        
        # Write transactions
        for transaction in transactions:
            writer.writerow([
                transaction.date.strftime('%Y-%m-%d'),
                transaction.amount,
                transaction.type,
                transaction.category.name,
                transaction.description
            ])
        
        # Create response
        output.seek(0)
        return Response(
            output.getvalue(),
            mimetype='text/csv',
            headers={
                'Content-Disposition': 'attachment; filename=transactions.csv'
            }
        )
        
    except Exception as e:
        current_app.logger.error(f"Error exporting CSV: {str(e)}")
        flash('Error exporting transactions', 'error')
        return redirect(url_for('account_settings'))

@app.route('/reset_data', methods=['POST'])
@login_required
def reset_data():
    try:
        # Delete all transactions for the user
        Transaction.query.filter_by(user_id=current_user.id).delete()
        db.session.commit()
        flash('All transaction data has been reset', 'success')
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error resetting data: {str(e)}")
        flash('Error resetting data', 'error')
    return redirect(url_for('account_settings'))


@app.route('/update_profile', methods=['POST'])
@login_required
def update_profile():
    try:
        # Get form data
        first_name = request.form.get('first_name', '').strip()
        last_name = request.form.get('last_name', '').strip()
        email = request.form.get('email', '').strip()
        
        # Validate input
        if not all([first_name, last_name, email]):
            flash('All fields are required', 'error')
            return redirect(url_for('account_settings'))
        
        # Update user information
        current_user.first_name = first_name
        current_user.last_name = last_name
        current_user.email = email
        
        db.session.commit()
        flash('Profile updated successfully!', 'success')
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error updating profile: {str(e)}")
        flash('An error occurred while updating profile', 'error')
    
    return redirect(url_for('account_settings'))

@app.route('/update_password', methods=['POST'])
@login_required
def update_password():
    try:
        current_password = request.form.get('current_password')
        new_password = request.form.get('new_password')
        confirm_password = request.form.get('confirm_password')
        
        # Validate input
        if not all([current_password, new_password, confirm_password]):
            flash('All password fields are required', 'error')
            return redirect(url_for('account_settings'))
        
        if new_password != confirm_password:
            flash('New passwords do not match', 'error')
            return redirect(url_for('account_settings'))
        
        if not current_user.check_password(current_password):
            flash('Current password is incorrect', 'error')
            return redirect(url_for('account_settings'))
        
        # Update password
        current_user.set_password(new_password)
        db.session.commit()
        flash('Password updated successfully!', 'success')
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error updating password: {str(e)}")
        flash('An error occurred while updating password', 'error')
    
    return redirect(url_for('account_settings'))

@app.route('/auth/register', methods=['GET', 'POST'])
def auth_register():
    if current_user.is_authenticated:
        return redirect(url_for('home'))
    
    if request.method == 'POST':
        try:
            email = request.form.get('email')
            password = request.form.get('password')
            first_name = request.form.get('first_name')
            last_name = request.form.get('last_name')

            # Check if user already exists
            if User.query.filter_by(email=email).first():
                flash('Email address already registered', 'error')
                return redirect(url_for('auth_register'))

            # Create and save the user first
            new_user = User(
                email=email,
                first_name=first_name,
                last_name=last_name,
            )
            new_user.set_password(password)
            
            try:
                # Add and commit the user first
                db.session.add(new_user)
                db.session.commit()  # Commit the user first
                
                # Define default categories
                default_categories = [
                    {'name': 'Salary', 'type': 'Income'},
                    {'name': 'Freelance', 'type': 'Income'},
                    {'name': 'Investments', 'type': 'Income'},
                    {'name': 'Groceries', 'type': 'Expense'},
                    {'name': 'Rent', 'type': 'Expense'},
                    {'name': 'Utilities', 'type': 'Expense'},
                    {'name': 'Transportation', 'type': 'Expense'},
                    {'name': 'Entertainment', 'type': 'Expense'}
                ]
                
                # Create categories for the new user
                for cat_data in default_categories:
                    try:
                        # Check if category already exists
                        existing = Category.query.filter_by(
                            user_id=new_user.id,
                            name=cat_data['name'],
                            type=cat_data['type']
                        ).first()
                        
                        if not existing:
                            category = Category(
                                name=cat_data['name'],
                                type=cat_data['type'],
                                user_id=new_user.id,
                                description=f"Default {cat_data['type'].lower()} category",
                                budget_limit=0.0
                            )
                            db.session.add(category)
                    except Exception as cat_error:
                        current_app.logger.error(f"Error adding category: {str(cat_error)}")
                        continue
                
                db.session.commit()
                
                # Log in the new user
                login_user(new_user)
                flash('Registration successful!', 'success')
                return redirect(url_for('home'))
                
            except Exception as e:
                db.session.rollback()
                current_app.logger.error(f"Database error during registration: {str(e)}", exc_info=True)
                flash('An error occurred during registration', 'error')
                return redirect(url_for('auth_register'))

        except Exception as e:
            current_app.logger.error(f"General registration error: {str(e)}", exc_info=True)
            flash('Registration error occurred', 'error')
            return redirect(url_for('auth_register'))

    return render_template('auth/register.html')







@app.route('/auth/login', methods=['GET', 'POST'])
def auth_login():
    if current_user.is_authenticated:
        return redirect(url_for('home'))

    if request.method == 'POST':
        try:
            email = request.form.get('email')
            password = request.form.get('password')
            remember = True if request.form.get('remember') else False

            user = User.query.filter_by(email=email).first()

            if user and user.check_password(password):
                db.session.commit()

                login_user(user, remember=remember)
                next_page = request.args.get('next')
                flash('Login successful!', 'success')        
                
                return redirect(next_page) if next_page else redirect(url_for('home'))
            else:
                flash('Invalid email or password', 'error')
                return redirect(url_for('auth_login'))

        except Exception as e:
            current_app.logger.error(f"Login error: {str(e)}", exc_info=True)
            flash(f'Login error: {str(e)}', 'error')
            return redirect(url_for('auth_login'))

    return render_template('auth/login.html')


@app.route('/auth/logout')
@login_required
def auth_logout():
    logout_user()
    flash('You have been logged out.', 'info')
    return redirect(url_for('auth_login'))


@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

@app.route('/expense-analysis', methods=['GET', 'POST'])
@login_required
def expense_analysis():
    try:
        # Get date range from query parameters or default to current month
        start_date_str = request.args.get('start_date')
        end_date_str = request.args.get('end_date')

        if start_date_str and end_date_str:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
        else:
            # Default to current month
            today = datetime.now()
            start_date = datetime(today.year, today.month, 1)
            end_date = (start_date + relativedelta(months=1))

        # Query all transactions within date range for current user
        transactions = Transaction.query.filter(
            Transaction.user_id == current_user.id,
            Transaction.date >= start_date,
            Transaction.date < end_date
        ).all()

        # Initialize counters and dictionaries
        total_income = total_expenses = 0
        income_by_category = {}
        expenses_by_category = {}

        # Process transactions
        for transaction in transactions:
            try:
                category_name = transaction.category.name
                amount = float(transaction.amount)

                # Check transaction type (case-insensitive)
                if transaction.type.lower() == 'income':
                    total_income += amount
                    income_by_category[category_name] = income_by_category.get(category_name, 0) + amount
                elif transaction.type.lower() == 'expense':
                    total_expenses += amount
                    expenses_by_category[category_name] = expenses_by_category.get(category_name, 0) + amount

            except Exception as e:
                current_app.logger.error(f"Error processing transaction {transaction.id}: {str(e)}")
                continue

        # Calculate category percentages
        income_percentages = []
        if total_income > 0:
            for category, amount in sorted(income_by_category.items(), key=lambda x: x[1], reverse=True):
                percentage = (amount / total_income) * 100
                income_percentages.append({
                    'category': category,
                    'amount': amount,
                    'percentage': round(percentage, 1)
                })

        expense_percentages = []
        if total_expenses > 0:
            for category, amount in sorted(expenses_by_category.items(), key=lambda x: x[1], reverse=True):
                percentage = (amount / total_expenses) * 100
                expense_percentages.append({
                    'category': category,
                    'amount': amount,
                    'percentage': round(percentage, 1)
                })

        # Get monthly trends for the last 12 months
        monthly_trends = []
        for i in range(11, -1, -1):
            month_start = start_date - relativedelta(months=i)
            month_end = month_start + relativedelta(months=1)

            # Get monthly totals
            month_transactions = Transaction.query.filter(
                Transaction.user_id == current_user.id,
                Transaction.date >= month_start,
                Transaction.date < month_end
            ).all()

            month_income = sum(t.amount for t in month_transactions if t.type.lower() == 'income')
            month_expenses = sum(t.amount for t in month_transactions if t.type.lower() == 'expense')
            
            monthly_trends.append({
                'month': month_start.strftime('%B %Y'),
                'income': month_income,
                'expenses': month_expenses,
                'net': month_income - month_expenses
            })

        # Get budget vs actual for expense categories
        budget_comparison = []
        expense_categories = Category.query.filter_by(
            user_id=current_user.id,
            type='expense'
        ).all()

        for category in expense_categories:
            spent = expenses_by_category.get(category.name, 0)
            if category.budget_limit > 0:
                percentage = (spent / category.budget_limit) * 100
            else:
                percentage = 0
                
            budget_comparison.append({
                'category': category.name,
                'budget': category.budget_limit,
                'spent': spent,
                'percentage': round(percentage, 1)
            })

        return render_template('expense_analysis.html',
                             start_date=start_date,
                             end_date=end_date,
                             total_income=total_income,
                             total_expenses=total_expenses,
                             net_income=total_income - total_expenses,
                             income_percentages=income_percentages,
                             expense_percentages=expense_percentages,
                             monthly_trends=monthly_trends,
                             budget_comparison=budget_comparison)

    except Exception as e:
        current_app.logger.error(f"Expense analysis error: {str(e)}")
        flash('An error occurred while generating the expense analysis.', 'error')
        return redirect(url_for('home'))

def check_system_categories():
    with app.app_context():
        system_cats = Category.query.filter_by(user_id=1).all()
        if not system_cats:
            print("No system categories found! Running init_db()...")
            init_db()
        else:
            print(f"Found {len(system_cats)} system categories:")
            for cat in system_cats:
                print(f"- {cat.name} ({cat.type})")

# Run this to check/initialize system categories
check_system_categories()


def export_settings():
    # Create a new workbook
    wb = Workbook()
    
    # Export Categories
    categories_sheet = wb.active
    categories_sheet.title = "Categories"
    categories_sheet.append(["ID", "Name", "Type"])
    categories = Category.query.all()
    for category in categories:
        categories_sheet.append([category.id, category.name, category.type])
    
    # Export Budget Goals
    budget_sheet = wb.create_sheet("BudgetGoals")
    budget_sheet.append(["Category_ID", "Amount", "Period"])
    budget_goals = BudgetGoal.query.all()
    for goal in budget_goals:
        budget_sheet.append([goal.category_id, goal.amount, goal.period])
    
    # Export Recurring Transactions
    recurring_sheet = wb.create_sheet("RecurringTransactions")
    recurring_sheet.append(["Type", "Category", "Amount", "Frequency", 
                          "Description", "Start_Date", "End_Date", "Is_Active"])
    recurring_transactions = RecurringTransaction.query.all()
    for trans in recurring_transactions:
        recurring_sheet.append([
            trans.type,
            trans.category,
            trans.amount,
            trans.frequency,
            trans.description,
            trans.start_date.strftime('%Y-%m-%d') if trans.start_date else None,
            trans.end_date.strftime('%Y-%m-%d') if trans.end_date else None,
            trans.is_active
        ])
    
    # Save to memory
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return excel_file

def import_settings(file_stream):
    try:
        wb = load_workbook(file_stream)
        
        # Import Categories
        if "Categories" in wb.sheetnames:
            categories_sheet = wb["Categories"]
            # Clear existing categories
            Category.query.delete()
            
            for row in list(categories_sheet.rows)[1:]:  # Skip header
                category = Category(
                    name=row[1].value,
                    type=row[2].value
                )
                db.session.add(category)
        
        # Import Budget Goals
        if "BudgetGoals" in wb.sheetnames:
            budget_sheet = wb["BudgetGoals"]
            # Clear existing budget goals
            BudgetGoal.query.delete()
            
            for row in list(budget_sheet.rows)[1:]:  # Skip header
                goal = BudgetGoal(
                    category_id=row[0].value,
                    amount=row[1].value,
                    period=row[2].value
                )
                db.session.add(goal)
        
        # Import Recurring Transactions
        if "RecurringTransactions" in wb.sheetnames:
            recurring_sheet = wb["RecurringTransactions"]
            # Clear existing recurring transactions
            RecurringTransaction.query.delete()
            
            for row in list(recurring_sheet.rows)[1:]:  # Skip header
                trans = RecurringTransaction(
                    type=row[0].value,
                    category=row[1].value,
                    amount=row[2].value,
                    frequency=row[3].value,
                    description=row[4].value,
                    start_date=datetime.strptime(row[5].value, '%Y-%m-%d') if row[5].value else None,
                    end_date=datetime.strptime(row[6].value, '%Y-%m-%d') if row[6].value else None,
                    is_active=row[7].value
                )
                db.session.add(trans)
        
        db.session.commit()
        return True, "Settings imported successfully"
    except Exception as e:
        db.session.rollback()
        return False, f"Error importing settings: {str(e)}"



def get_monthly_trends(months):
    try:
        # Calculate date range
        end_date = datetime.now()
        start_date = datetime(end_date.year - 1, end_date.month + 1, 1) if end_date.month < 12 else datetime(end_date.year, 1, 1)

        # Query all transactions within date range
        transactions = Transaction.query.filter(
            Transaction.date >= start_date,
            Transaction.date <= end_date
        ).all()

        # Initialize monthly totals
        monthly_totals = {}
        current_date = start_date
        while current_date <= end_date:
            month_key = current_date.strftime('%Y-%m')
            monthly_totals[month_key] = {'income': 0, 'expenses': 0, 'net': 0}
            current_date = (current_date.replace(day=28) + timedelta(days=4)).replace(day=1)

        # Calculate monthly totals
        for transaction in transactions:
            month_key = transaction.date.strftime('%Y-%m')
            if transaction.type == 'income':
                monthly_totals[month_key]['income'] += transaction.amount
            else:
                monthly_totals[month_key]['expenses'] += transaction.amount
            monthly_totals[month_key]['net'] = (
                monthly_totals[month_key]['income'] - 
                monthly_totals[month_key]['expenses']
            )

        # Convert to list of tuples sorted by date
        trends = []
        for month_key in sorted(monthly_totals.keys()):
            month_data = monthly_totals[month_key]
            month_label = datetime.strptime(month_key, '%Y-%m').strftime('%b %Y')
            trends.append((
                month_label,
                month_data['income'],
                month_data['expenses'],
                month_data['net']
            ))

        return trends

    except Exception as e:
        print(f"Error calculating monthly trends: {str(e)}")
        return []


@app.route('/income_analysis')
def income_analysis():
    incomes = Transaction.query.filter_by(type='Income').all()
    if not incomes:
        flash('No transactions available for analysis.', 'info')
        return render_template('income_analysis.html')
    
    # Prepare data for charts
    income_by_category = {}
    for income in incomes:
        if income.category in income_by_category:
            income_by_category[income.category] += income.amount
        else:
            income_by_category[income.category] = income.amount
    
    # Create pie chart
    fig_pie = px.pie(
        values=list(income_by_category.values()),
        names=list(income_by_category.keys()),
        title='Income by Category'
    )
    
    # Create time series chart
    incomes_by_date = {}
    for income in incomes:
        month = income.date.strftime('%Y-%m')
        if month in incomes_by_date:
            incomes_by_date[month] += income.amount
        else:
            incomes_by_date[month] = income.amount
    
    sorted_dates = sorted(incomes_by_date.keys())
    fig_line = px.line(
        x=sorted_dates,
        y=[incomes_by_date[date] for date in sorted_dates],
        title='Monthly Income Over Time'
    )
    
    charts = {
        'pie': json.dumps(fig_pie, cls=plotly.utils.PlotlyJSONEncoder),
        'line': json.dumps(fig_line, cls=plotly.utils.PlotlyJSONEncoder)
    }
    
    return render_template('income_analysis.html',
                         charts=charts,
                         income_data=income_by_category)
@app.route('/get_transaction/<int:transaction_id>')
def get_transaction(transaction_id):
    try:
        transaction = Transaction.query.get_or_404(transaction_id)
        return jsonify({
            'date': transaction.date.strftime('%Y-%m-%d'),
            'type': transaction.type,
            'category': transaction.category,
            'amount': float(transaction.amount),  # Ensure it's a float
            'description': transaction.description or ''  # Handle None values
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/edit_transaction', methods=['POST'])
@login_required
def edit_transaction():
    try:
        transaction_id = request.form.get('transaction_id')
        transaction = Transaction.query.get_or_404(transaction_id)

        # Verify ownership
        if transaction.user_id != current_user.id:
            flash('Unauthorized access to transaction', 'error')
            return redirect(url_for('view_transactions'))

        # Update transaction details
        transaction.date = datetime.strptime(request.form.get('date'), '%Y-%m-%d')
        transaction.type = request.form.get('type')
        transaction.category_id = request.form.get('category_id')
        transaction.amount = float(request.form.get('amount'))
        transaction.description = request.form.get('description')

        db.session.commit()
        flash('Transaction updated successfully!', 'success')

    except ValueError:
        flash('Invalid amount or date format!', 'error')
    except Exception as e:
        db.session.rollback()
        flash(f'Error updating transaction: {str(e)}', 'error')

    return redirect(url_for('view_transactions'))


@app.route('/delete_transaction/<int:transaction_id>', methods=['POST'])
def delete_transaction(transaction_id):
    try:
        transaction = Transaction.query.get_or_404(transaction_id)
        db.session.delete(transaction)
        db.session.commit()
        flash('Transaction deleted successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting transaction: {str(e)}', 'error')
    return redirect(url_for('view_transactions'))

scheduler = BackgroundScheduler()
scheduler.add_job(
    func=process_recurring_transactions,
    trigger=CronTrigger(hour=0, minute=0),  # Run at midnight every day
    id='process_recurring_transactions',
    name='Process recurring transactions',
    replace_existing=True
)
scheduler.start()




# Error handlers
@app.errorhandler(404)
def page_not_found(e):
    return render_template('404.html'), 404

@app.errorhandler(500)
def internal_server_error(e):
    return render_template('500.html'), 500

if __name__ == '__main__':
    with app.app_context():
        init_db()  # Initialize database and add default categories
        db.create_all()
    app.run(debug=True)

